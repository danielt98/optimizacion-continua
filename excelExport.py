import openpyxl

# creating the DataFrame
class excelExport:
    def __init__(self, data):
        self.data = data

    def evaluate(self):
        wb = openpyxl.Workbook()
        hoja = wb.active
        hoja.title = "Continua"
        hoja.append(('Optiomizacion continua', 'Ascenso de la colina con remplazo','','','','','Optiomizacion continua opuesta por caos'))
        hoja.append(('', 'Media','Desviación','Mejor','Peor','Tiempo','Media','Desviación','Mejor','Peor','Tiempo'))
        hoja.merge_cells(start_row=1, start_column=2, end_row=1, end_column=6)
        hoja.merge_cells(start_row=1, start_column=7, end_row=1, end_column=11)
        hoja.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        for function in self.data:
            hoja.append((function, self.data[function][1][0].mean(), self.data[function][1][0].std(), self.data[function][1][0].max(), self.data[function][1][0].min(),self.data[function][2][0].min(),
                                   self.data[function][1][1].mean(), self.data[function][1][1].std(), self.data[function][1][1].max(), self.data[function][1][1].min(), self.data[function][2][1].min()))
        wb.save('opContinua.xlsx')
